from decimal import Decimal
from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, send_file
from utils.db import get_connection
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from router.accesos import requiere_acceso

# Definir el Blueprint
ventas_zonas_mes_bp = Blueprint("ventas_zonas_mes", __name__)

def obtener_reporte_ventas_zonas_mes(fecha_inicio, fecha_fin, vendedor=None, departamento=None, municipio=None):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    start_date = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    end_date = datetime.strptime(fecha_fin, "%Y-%m-%d")

    meses_rango = []
    fecha_actual = start_date
    while fecha_actual <= end_date:
        nombre_mes = fecha_actual.strftime("%b_%Y")  # Ej: Jul_2026
        meses_rango.append(nombre_mes)
        if fecha_actual.month == 12:
            fecha_actual = datetime(fecha_actual.year + 1, 1, 1)
        else:
            fecha_actual = datetime(fecha_actual.year, fecha_actual.month + 1, 1)

    # columnas dinámicas
    select_parts = [
        "e.tipo AS TipoVendedor",
        "d.vendedor",
        "z.nombre_zona",
        "(CASE WHEN c.direccion IS NULL THEN 'OCACIONAL' ELSE z.nombre_zona END) AS ZONA",
        "UPPER(c.departamento) AS Departamento",
        "UPPER(c.municipio) AS Municipio",
        "d.Cliente",
        "c.direccion",
        "d.Marca"
    ]

    for nombre_mes in meses_rango:
        dt = datetime.strptime(nombre_mes, "%b_%Y")
        inicio_mes = dt.replace(day=1).strftime("%Y-%m-%d")
        if dt.month == 12:
            fin_mes = dt.replace(year=dt.year+1, month=1, day=1) - timedelta(days=1)
        else:
            fin_mes = dt.replace(month=dt.month+1, day=1) - timedelta(days=1)
        fin_mes = fin_mes.strftime("%Y-%m-%d")

        select_parts.append(
            f"SUM(CASE WHEN d.fecha BETWEEN '{inicio_mes}' AND '{fin_mes}' "
            f"THEN d.`Sub-Total 2` ELSE 0 END) AS `{nombre_mes}`"
        )

    # Totales
    select_parts.append("SUM(d.`Sub-Total 2`) AS TotalRango")
    select_parts.append("tg.TotalGeneral")

    query = f"""
        SELECT {", ".join(select_parts)}
        FROM ventas_por_producto d
        LEFT JOIN clientes c ON d.Cliente=c.nombre
        LEFT JOIN zona_cliente z ON c.municipio=z.departamento
        LEFT JOIN (SELECT Cliente, SUM(`Sub-Total 2`) AS TotalGeneral
                   FROM ventas_por_producto
                   WHERE Estatus!='Anulado'
                   GROUP BY Cliente) tg ON d.Cliente = tg.Cliente
        LEFT JOIN (SELECT tipo, nombre FROM cat_vendedor GROUP BY tipo, nombre) e ON d.Vendedor = e.nombre
        WHERE d.fecha BETWEEN %s AND %s AND Estatus!='Anulado'
        {"AND d.vendedor=%s" if vendedor else ""}
        {"AND UPPER(c.departamento)=%s" if departamento else ""}
        {"AND UPPER(c.municipio)=%s" if municipio else ""}
        GROUP BY e.tipo,d.vendedor,d.Cliente,c.departamento,c.municipio,z.nombre_zona,c.direccion,d.Marca,tg.TotalGeneral
        ORDER BY z.nombre_zona,c.departamento,c.municipio;
    """

    params = [fecha_inicio, fecha_fin]
    if vendedor:
        params.append(vendedor)
    if departamento:
        params.append(departamento.upper())
    if municipio:
        params.append(municipio.upper())

    cursor.execute(query, tuple(params))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    return {
        "registros": rows,
        "meses_rango": meses_rango,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "vendedor": vendedor,
        "departamento": departamento,
        "municipio": municipio
    }

def obtener_vendedores():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT nombre FROM cat_vendedor GROUP BY nombre ORDER BY nombre;")
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return [row["nombre"] for row in rows]

def obtener_departamentos():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT DISTINCT UPPER(departamento) AS departamento FROM clientes ORDER BY departamento;")
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return [row["departamento"] for row in rows if row["departamento"]]

def obtener_municipios():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT DISTINCT UPPER(municipio) AS municipio FROM clientes ORDER BY municipio;")
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return [row["municipio"] for row in rows if row["municipio"]]

@ventas_zonas_mes_bp.route("/ventas_zonas_mes", methods=["GET", "POST"])
@requiere_acceso("ventas_zonas_mes")
def reporte_ventas_zonas_mes():
    if request.method == "POST":
        fecha_inicio = request.form.get("fecha_inicio", datetime.today().strftime("%Y-%m-%d"))
        fecha_fin = request.form.get("fecha_fin", datetime.today().strftime("%Y-%m-%d"))
        vendedor = request.form.get("vendedor")
        departamento = request.form.get("departamento")
        municipio = request.form.get("municipio")
    else:
        fecha_inicio = request.args.get("fecha_inicio", datetime.today().strftime("%Y-%m-%d"))
        fecha_fin = request.args.get("fecha_fin", datetime.today().strftime("%Y-%m-%d"))
        vendedor = request.args.get("vendedor")
        departamento = request.args.get("departamento")
        municipio = request.args.get("municipio")

    datos = obtener_reporte_ventas_zonas_mes(fecha_inicio, fecha_fin, vendedor, departamento, municipio)
    vendedores = obtener_vendedores()
    departamentos = obtener_departamentos()
    municipios = obtener_municipios()

    return render_template("reporte_ventas_zonas_mes.html",
                           **datos,
                           vendedores=vendedores,
                           departamentos=departamentos,
                           municipios=municipios)

# Exportar en Excel con formato
@ventas_zonas_mes_bp.route("/exportar_reporte_ventas_zonas_mes")
def exportar_reporte_ventas_zonas_mes():
    fecha_inicio = request.args.get("fecha_inicio", datetime.today().strftime("%Y-%m-%d"))
    fecha_fin = request.args.get("fecha_fin", datetime.today().strftime("%Y-%m-%d"))
    vendedor = request.args.get("vendedor")
    departamento = request.args.get("departamento")
    municipio = request.args.get("municipio")

    datos = obtener_reporte_ventas_zonas_mes(fecha_inicio, fecha_fin, vendedor, departamento, municipio)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Ventas"

    headers = ["Tipo Vendedor", "Vendedor", "Cliente", "Zona", "Departamento", "Municipio", "Marca"]
    headers.extend(datos["meses_rango"])
    headers.extend(["TotalRango", "TotalGeneral"])
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in datos["registros"]:
        fila = [
            r["TipoVendedor"],
            r["vendedor"],
            r["Cliente"],
            r["ZONA"],
            r["Departamento"],
            r["Municipio"],
            r["Marca"]
        ]
        for mes in datos["meses_rango"]:
            fila.append(r[mes] if r[mes] is not None else 0)
        fila.append(r["TotalRango"])
        fila.append(r["TotalGeneral"])
        ws.append(fila)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"reporte_ventas_{fecha_inicio}_a_{fecha_fin}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


